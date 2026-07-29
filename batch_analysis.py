#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2025-2026 Sergei Sergeev
# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Commercial use requires a separate license: see COMMERCIAL-LICENSE.md
"""
IVF Digital Twin v6.2 -- Batch Analysis
=========================================================
Запуск:
    python batch_analysis.py <input.xlsx> [output_dir] [--clinic "Название клиники"]

Примеры:
    python batch_analysis.py февраль.xlsx
    python batch_analysis.py февраль.xlsx results/ --clinic "Клиника А"

Что делает:
    1. Читает Excel-выгрузку клиники (формат «отчет.xlsx»).
    2. Для каждой пациентки запускает predict_single_patient() из src/ivf_core.py
       — полный L1-L6: ZINB-воронка, FORTUNE+KPI, KAT-нейросеть (если веса есть),
       байес-posterior, кластер L4, CSDI Hybrid v3 (L5), GNN L6.
    3. Дописывает строки в dt_analytics_data/dt_predictions.csv (схема app.py + GNN).
    4. Создаёт OPU_table_filled.xlsx (шаблон OPU_table с прогнозом DT).
    5. Создаёт dt_vs_real_comparison.csv (прогноз DT vs факт по воронке).
    6. Создаёт batch_summary.csv — сводка по всем пациенткам со всеми слоями.

Что подаётся в модель (4 «формных» значения + метаданные):
    Возраст               → age           (обязательно)
    Количество фолликулов → follicles      (если есть)
    Число ОКК             → known.okk     (если есть)
    Число инсеминированных → known.mii    (по указанию автора модели)
    № попытки             → attempt
    АМГ / КАФ             → см. USE_REPORT_AMH_AFC
    BMI                   → дефолт 23.0

Downstream (2pN, бластоцисты и т.д.) в модель НЕ подаются —
они прогнозируются и сравниваются с фактом.
"""

import sys, os, argparse, csv, re
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd
# Windows console encoding fix
import sys as _sys
if hasattr(_sys.stdout, 'reconfigure'):
    try:
        _sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import warnings
warnings.filterwarnings("ignore", message="X does not have valid feature names")

# ── Пути к модулям ────────────────────────────────────────────────────────────
_HERE     = os.path.dirname(os.path.abspath(__file__))
_SRC_DIR  = os.path.join(_HERE, "src")
sys.path.insert(0, _SRC_DIR)
sys.path.insert(0, _HERE)

from ivf_core import (  # noqa: E402
    predict_single_patient,
    save_analytics_record,
    _ANALYTICS_COLUMNS,
    _BASE_DIR,
)
from befe_batch_utils import compute_l7_posterior  # noqa: E402

# ══════════════════════════════════════════════════════════════════════════════
#  КОНФИГУРАЦИЯ — меняйте под клинику
# ══════════════════════════════════════════════════════════════════════════════
N_SIM               = 2000     # Monte Carlo iterations на пациентку
SEED_BASE           = 42       # базовый seed (к нему прибавляется номер строки)
DEFAULT_BMI         = 23.0
DEFAULT_AMH         = 2.0      # если АМГ отсутствует или некорректен
DEFAULT_AFC         = 12       # если КАФ отсутствует или некорректен
USE_REPORT_AMH      = False    # True → брать АМГ из выгрузки (с проверкой)
USE_REPORT_AFC      = False    # True → брать КАФ из выгрузки (с проверкой)
AMH_PLAUSIBLE       = (0.01, 30.0)
AFC_PLAUSIBLE       = (1, 60)
SPERM_SOURCE        = "ejaculate"
# Реальные клинические батчи для байес-prior (None → covariate-dependent prior):
CLINIC_SUCCESSES    = None     # напр. [19, 18, 20]
CLINIC_TRIALS       = None     # напр. [43, 45, 65]
DT_HEADLINE_COL     = "p_overall_cycle"  # fallback для колонки «DIGITAL TWIN»

# ── Маппинг колонок входного файла (формат «отчет.xlsx» клиники) ─────────────
COL = {
    "fio":          "ФИО",
    "dob":          "Дата рождения",
    "age":          "Возраст",
    "id":           "Номер карты пациента",
    "attempt":      "№ попытки",
    "amh":          "АМГ",
    "afc":          "КАФ",
    "foll":         "Количество фолликулов",
    "okk":          "Число ОКК",
    "mii":          "Число MII",
    "insem":        "Число инсеминированных",
    "pn2":          "2 pN",
    "cleav":        "Число дробящихся на 3 день",
    "bl":           "Число Bl",
    "goodbl":       "Число Bl хор.кач-ва",
    "cryo":         "Заморожено бластоцист",
    "date_opu":     "Дата пункции",
    "date_et":      "Дата переноса",
    "day_et":       "День переноса",
    "et_n":         "Перенесено эмбрионов",
    "outcome":      "Исход переноса",
}

# OPU_table выходные колонки
OPU_COLUMNS = [
    "Patient Full Name", "DOB", "ID", "Date OPU", "Date ET", "BMI", "AMH",
    "Attempt", "AFC", "Age", "N folicules OPU", "OCC", "MII", "Inseminated",
    "2pN", "Cleavage", "Bl", "Good Bl", "Cryo", "ET", "Day of ET", "Preg",
    "PRAI", "DIGITAL TWIN",
]

# ══════════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def parse_age(v) -> float | None:
    """Парсит возраст: int, float или строку '26 лет', '33 года', '29 years'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            x = float(v)
            return x if 10 < x < 60 else None
        except (ValueError, TypeError):
            return None
    s = str(v).strip()
    # Ищем первое число в строке
    m = re.search(r"(\d+)", s)
    if m:
        x = float(m.group(1))
        return x if 10 < x < 60 else None
    return None


def num(v, cast=float):
    """Безопасное приведение ячейки к числу. Пустое/мусор → None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
        if v in ("", "-", "—", "н/д", "NA", "nan", "None"):
            return None
    try:
        x = float(v)
        if np.isnan(x):
            return None
        return int(round(x)) if cast is int else x
    except (ValueError, TypeError):
        return None


def opt_int(v) -> int | None:
    """Целое > 0. 0 и None → None (как optional_int в app.py)."""
    iv = num(v, int)
    return iv if (iv is not None and iv > 0) else None


def get_col(rec: dict, key: str):
    """Достаёт значение по символьному ключу из маппинга COL."""
    return rec.get(COL.get(key, ""))


def resolve_amh_afc(rec: dict):
    amh, afc = DEFAULT_AMH, DEFAULT_AFC
    if USE_REPORT_AMH:
        a = num(get_col(rec, "amh"), float)
        if a is not None and AMH_PLAUSIBLE[0] <= a <= AMH_PLAUSIBLE[1]:
            amh = a
    if USE_REPORT_AFC:
        f = num(get_col(rec, "afc"), int)
        if f is not None and AFC_PLAUSIBLE[0] <= f <= AFC_PLAUSIBLE[1]:
            afc = f
    return amh, afc


def build_data_quality_flags(rec: dict, age, okk, insem, foll) -> str:
    """Возвращает строку флагов качества данных."""
    flags = []
    # MII из поля MII (реальное значение для сравнения)
    real_mii = num(get_col(rec, "mii"), int)
    if okk is not None and real_mii is not None and real_mii > okk:
        flags.append(f"MII({real_mii})>OKK({okk})")
    if foll is None:
        flags.append("no_foll")
    if okk is None:
        flags.append("no_okk")
    afc_raw = num(get_col(rec, "afc"), int)
    if afc_raw is not None and afc_raw > AFC_PLAUSIBLE[1]:
        flags.append(f"AFC_implausible({afc_raw})")
    amh_raw = num(get_col(rec, "amh"), float)
    if amh_raw is not None and not (AMH_PLAUSIBLE[0] <= amh_raw <= AMH_PLAUSIBLE[1]):
        flags.append(f"AMH_implausible({amh_raw})")
    return ";".join(flags) if flags else "ok"


# ══════════════════════════════════════════════════════════════════════════════
#  ОБРАБОТКА ОДНОЙ СТРОКИ
# ══════════════════════════════════════════════════════════════════════════════

def process_row(rec: dict, row_index: int) -> tuple:
    """
    Возвращает (summary_row, opu_row, cmp_row, record_id) или None если нет возраста.
    """
    age = parse_age(get_col(rec, "age"))
    if age is None:
        return None

    amh, afc = resolve_amh_afc(rec)
    attempt   = opt_int(get_col(rec, "attempt")) or 1
    foll      = opt_int(get_col(rec, "foll"))
    okk       = opt_int(get_col(rec, "okk"))
    insem     = opt_int(get_col(rec, "insem"))   # → known.mii

    dq_flags  = build_data_quality_flags(rec, age, okk, insem, foll)

    # Реальные downstream-значения (для сравнения, в модель не идут)
    real_pn2    = opt_int(get_col(rec, "pn2"))
    real_cleav  = opt_int(get_col(rec, "cleav"))
    real_bl     = opt_int(get_col(rec, "bl"))
    real_goodbl = opt_int(get_col(rec, "goodbl"))
    real_cryo   = opt_int(get_col(rec, "cryo"))
    real_mii    = opt_int(get_col(rec, "mii"))   # реальное MII (для OPU-таблицы)
    real_outcome= str(get_col(rec, "outcome") or "").strip()

    # Нормализуем исход для PREG колонки
    preg_flag = ""
    if "беременность +" in real_outcome.lower() or real_outcome.lower() == "беременность":
        preg_flag = "Да"
    elif "не наступила" in real_outcome.lower():
        preg_flag = "Нет"

    # Запускаем predict_single_patient
    result = predict_single_patient(
        age=age, amh=amh, afc=afc, bmi=DEFAULT_BMI,
        attempt=attempt, follicles=foll,
        sperm_source=SPERM_SOURCE,
        known_okk=okk, known_mii=insem,
        clinic_successes=CLINIC_SUCCESSES,
        clinic_trials=CLINIC_TRIALS,
        n_sim=N_SIM,
        seed=SEED_BASE + row_index,
    )
    res  = result["res"]
    post = res.get("posterior", {})
    ca   = res.get("cluster_analysis", {})
    probs= ca.get("cluster_probs", {})

    p_kat    = result.get("p_kat_raw")
    p_nvsa   = result.get("p_nvsa")
    p_csdi   = result.get("p_csdi")
    p_gnn_ens= result.get("p_gnn_ens")
    p_l7, _, _ = compute_l7_posterior(
        result,
        age=age, amh=amh, afc=afc, bmi=DEFAULT_BMI,
        base_dir=_BASE_DIR,
    )
    headline = p_l7 if p_l7 is not None else res.get(DT_HEADLINE_COL, res.get("p_overall_cycle", 0))

    r4 = lambda v: round(float(v), 4) if v is not None else None

    # ── Сводная строка batch_summary.csv ──────────────────────────────────
    summary = {
        "row":           row_index + 1,
        "patient_name":  get_col(rec, "fio") or "",
        "patient_id":    get_col(rec, "id") or "",
        "age":           age,
        "amh_used":      amh,
        "afc_used":      afc,
        "attempt":       attempt,
        "follicles":     foll if foll is not None else "",
        "known_okk":     okk  if okk  is not None else "",
        "known_mii":     insem if insem is not None else "",
        "data_quality":  dq_flags,
        # Медианы воронки
        "med_okk":       res["okk_med"],
        "med_mii":       res["mii_med"],
        "med_pn2":       res["pn2_med"],
        "med_blasts":    res["blasts_med"],
        "med_good":      res["good_med"],
        # Прогнозы беременности (все слои)
        "p_per_transfer":    r4(res.get("p_per_transfer")),
        "p_overall_cycle":   r4(res.get("p_overall_cycle")),
        "bayes_mean":        r4(post.get("mean")),
        "p_kat_raw":         r4(p_kat),
        "p_nvsa":            r4(p_nvsa),
        "p_csdi":            r4(p_csdi),
        "p_gnn_ens":         r4(p_gnn_ens),
        # Кластер L4
        "dominant_cluster":  ca.get("dominant_cluster", ""),
        "c0_prob":           r4(probs.get(0)),
        "c1_prob":           r4(probs.get(1)),
        "c2_prob":           r4(probs.get(2)),
        # Риски
        "ohss_any":          r4(res.get("ohss", {}).get("p_any_ohss", 0)),
        "p_cancel":          r4(float(np.mean(res["sim_okk"] == 0))),
        # Слои доступности
        "nn_available":      int(result["nn_available"]),
        "csdi_available":    int(result["csdi_available"]),
        "gnn_available":     int(result["gnn_available"]),
        "nn_source":         res.get("nn_prediction", {}).get("source", ""),
        # Реальные данные
        "real_pn2":          real_pn2    if real_pn2    is not None else "",
        "real_bl":           real_bl     if real_bl     is not None else "",
        "real_goodbl":       real_goodbl if real_goodbl is not None else "",
        "real_outcome":      real_outcome,
    }

    # ── OPU_table строка ──────────────────────────────────────────────────
    opu = {
        "Patient Full Name": get_col(rec, "fio") or "",
        "DOB":               get_col(rec, "dob") or "",
        "ID":                get_col(rec, "id") or "",
        "Date OPU":          get_col(rec, "date_opu") or "",
        "Date ET":           get_col(rec, "date_et") or "",
        "BMI":               DEFAULT_BMI,
        "AMH":               num(get_col(rec, "amh"), float) or "",
        "Attempt":           attempt,
        "AFC":               num(get_col(rec, "afc"), int) or "",
        "Age":               age,
        "N folicules OPU":   foll if foll is not None else "",
        "OCC":               okk  if okk  is not None else "",
        "MII":               real_mii if real_mii is not None else "",
        "Inseminated":       insem if insem is not None else "",
        "2pN":               real_pn2    if real_pn2    is not None else "",
        "Cleavage":          real_cleav  if real_cleav  is not None else "",
        "Bl":                real_bl     if real_bl     is not None else "",
        "Good Bl":           real_goodbl if real_goodbl is not None else "",
        "Cryo":              real_cryo   if real_cryo   is not None else "",
        "ET":                num(get_col(rec, "et_n"), int) or "",
        "Day of ET":         num(get_col(rec, "day_et"), int) or "",
        "Preg":              preg_flag,
        "PRAI":              "",
        "DIGITAL TWIN":      r4(headline),
    }

    # ── Comparison row (прогноз DT vs факт) ──────────────────────────────
    cmp = {
        "patient_name":       get_col(rec, "fio") or "",
        "patient_id":         get_col(rec, "id") or "",
        "age":                age,
        "attempt":            attempt,
        "follicles_in":       foll,
        "okk_in":             okk,
        "insem_in":           insem,
        # Факт
        "real_pn2":           real_pn2,
        "real_bl":            real_bl,
        "real_goodbl":        real_goodbl,
        "real_outcome":       real_outcome,
        # Прогноз DT (медианы воронки)
        "dt_pn2":             res["pn2_med"],
        "dt_bl":              res["blasts_med"],
        "dt_goodbl":          res["good_med"],
        # Вероятности беременности — все слои
        "p_per_transfer":     r4(res.get("p_per_transfer")),
        "p_overall_cycle":    r4(res.get("p_overall_cycle")),
        "bayes_mean":         r4(post.get("mean")),
        "p_kat_raw":          r4(p_kat),
        "p_nvsa":             r4(p_nvsa),
        "p_csdi":             r4(p_csdi),
        "p_gnn_ens":          r4(p_gnn_ens),
        "dominant_cluster":   ca.get("dominant_cluster", ""),
        "data_quality":       dq_flags,
    }

    return summary, opu, cmp, result


# ══════════════════════════════════════════════════════════════════════════════
#  ЧТЕНИЕ EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def read_records(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    recs = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        recs.append({header[i]: r[i]
                     for i in range(min(len(header), len(r)))})
    return recs


# ══════════════════════════════════════════════════════════════════════════════
#  ЗАПИСЬ OPU-ТАБЛИЦЫ С ФОРМАТИРОВАНИЕМ
# ══════════════════════════════════════════════════════════════════════════════

def write_opu_xlsx(opu_rows: list, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import openpyxl.utils as oxl_utils
    wb = Workbook()
    ws = wb.active
    ws.title = "OPU_DT"
    ws.append(OPU_COLUMNS)
    hfont = Font(name="Arial", bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1B4F72")
    dtfill = PatternFill("solid", start_color="FCE4D6")
    for col in range(1, len(OPU_COLUMNS) + 1):
        c = ws.cell(row=1, column=col)
        c.font  = hfont
        c.fill  = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for orow in opu_rows:
        ws.append([orow.get(k, "") for k in OPU_COLUMNS])
    dt_col = OPU_COLUMNS.index("DIGITAL TWIN") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=dt_col).fill = dtfill
    for col in range(1, len(OPU_COLUMNS) + 1):
        ws.column_dimensions[oxl_utils.get_column_letter(col)].width = 14
    ws.freeze_panes = "A2"
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="IVF Digital Twin — пакетный анализ выгрузки клиники"
    )
    parser.add_argument("input",  help="Путь к Excel-выгрузке клиники")
    parser.add_argument("output", nargs="?", default=".",
                        help="Каталог для результатов (по умолчанию '.')")
    parser.add_argument("--clinic", default="",
                        help="Название клиники для записи в CSV")
    args = parser.parse_args()

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Путь к dt_predictions.csv (пишем в dt_analytics_data внутри проекта,
    # точно как app.py — чтобы строки объединялись с интерактивными)
    analytics_csv = str(Path(_BASE_DIR) / "dt_analytics_data" / "dt_predictions.csv")

    print(f"\n{'='*60}")
    print(f" IVF Digital Twin v6.2 -- Batch Analysis")
    print(f" Input:   {args.input}")
    print(f" Clinic:  {args.clinic or '(not specified)'}")
    print(f"{'='*60}\n")

    recs = read_records(args.input)
    print(f"Records loaded: {len(recs)}")

    summaries, opu_rows, cmp_rows = [], [], []
    skipped, errors = 0, 0

    for i, rec in enumerate(recs):
        fio = rec.get(COL["fio"]) or f"строка {i+2}"
        try:
            result_tuple = process_row(rec, i)
            if result_tuple is None:
                print(f"  [{i+1:03d}] SKIPPED (no age): {fio}")
                skipped += 1
                continue

            summary, opu, cmp, full_result = result_tuple
            summaries.append(summary)
            opu_rows.append(opu)
            cmp_rows.append(cmp)

            # Дописываем в dt_predictions.csv (как app.py при генерации PDF)
            record_id = save_analytics_record(
                result       = full_result,
                age          = summary["age"],
                amh          = summary["amh_used"],
                afc          = summary["afc_used"],
                bmi          = DEFAULT_BMI,
                attempt      = summary["attempt"],
                sperm_source = SPERM_SOURCE,
                follicles    = summary["follicles"] if summary["follicles"] != "" else None,
                clinic_name  = args.clinic or "",
                patient_name = summary["patient_name"],
                patient_id   = summary["patient_id"],
                real_pn2     = summary["real_pn2"],
                real_cleav   = "",
                real_bl      = summary["real_bl"],
                real_goodbl  = summary["real_goodbl"],
                real_cryo    = "",
                real_outcome = summary["real_outcome"],
                notes        = f"batch;dq={summary['data_quality']}",
                analytics_csv= analytics_csv,
            )
            summary["record_id"] = record_id or ""

            # Прогресс каждые 10 пациенток
            nn_ok  = "✓NN" if summary["nn_available"]  else " NN-"
            cs_ok  = "✓CS" if summary["csdi_available"] else " CS-"
            gn_ok  = "✓GN" if summary["gnn_available"]  else " GN-"
            print(f"  [{i+1:03d}] {fio[:40]:<40}  "
                  f"age={summary['age']:.0f}  "
                  f"p_overall={summary['p_overall_cycle'] or '—'}  "
                  f"{nn_ok}{cs_ok}{gn_ok}  DQ={summary['data_quality']}")

        except Exception as exc:
            print(f"  [{i+1:03d}] ERROR: {fio} | {exc}")
            errors += 1

    print(f"\n{'='*60}")
    print(f"  OK: {len(summaries)}  "
          f"Skipped (no age): {skipped}  "
          f"Errors: {errors}")
    print(f"{'='*60}\n")

    # ── Записываем файлы ──────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # 1. batch_summary.csv
    summ_path = str(out_dir / f"batch_summary_{ts}.csv")
    pd.DataFrame(summaries).to_csv(summ_path, index=False, encoding="utf-8-sig")
    print(f"[OK] batch_summary       -> {summ_path}")

    # 2. OPU_table_filled.xlsx
    opu_path = str(out_dir / f"OPU_table_filled_{ts}.xlsx")
    write_opu_xlsx(opu_rows, opu_path)
    print(f"[OK] OPU_table_filled    -> {opu_path}")

    # 3. dt_vs_real_comparison.csv
    cmp_path = str(out_dir / f"dt_vs_real_{ts}.csv")
    pd.DataFrame(cmp_rows).to_csv(cmp_path, index=False, encoding="utf-8-sig")
    print(f"[OK] dt_vs_real          -> {cmp_path}")

    # 4. dt_predictions.csv (сообщаем путь)
    print(f"[OK] dt_predictions.csv  -> {analytics_csv}")
    print()


if __name__ == "__main__":
    main()
