#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2025-2026 Sergei Sergeev
# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Commercial use requires a separate license: see COMMERCIAL-LICENSE.md
"""
IVF Digital Twin -- Post-processing
====================================
Reads the predictions Excel (report + Predictions column),
runs full DT pipeline L1-L6 for each patient, writes:
  - dt_analytics_data/dt_predictions.csv   (66 cols, same as app.py)
  - results/OPU_table_filled_*.xlsx
  - results/dt_analytics_ready_*.csv       (66 cols + PRAI + Preg binary)
  - results/dt_analytics_with_outcome_*.csv (rows with known outcome only)

Usage:
    python dt_postprocess.py <predictions_file.xlsx> [output_dir] [--clinic NAME]
"""

import sys, os, warnings
warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd

_HERE    = os.path.dirname(os.path.abspath(__file__))
_SRC_DIR = os.path.join(_HERE, "src")
sys.path.insert(0, _SRC_DIR)
sys.path.insert(0, _HERE)

from befe_batch_utils import compute_l7_posterior

# ── Column names in the predictions file ──────────────────────────────────────
COL_ID       = "Номер карты пациента"
COL_FIO      = "ФИО"
COL_AGE      = "Возраст"
COL_DOB      = "Дата рождения"
COL_ATTEMPT  = "№ попытки"
COL_AMH      = "АМГ"
COL_AFC      = "КАФ"
COL_FOLL     = "Количество фолликулов"
COL_OKK      = "Число ОКК"
COL_MII      = "Число MII"
COL_INSEM    = "Число инсеминированных"
COL_PN2      = "2 pN"
COL_CLEAV    = "Число дробящихся на 3 день"
COL_BL       = "Число Bl"
COL_GOODBL   = "Число Bl хор.кач-ва"
COL_CRYO     = "Заморожено бластоцист"
COL_DATE_OPU = "Дата пункции"
COL_DATE_ET  = "Дата переноса"
COL_DAY_ET   = "День переноса"
COL_ET_N     = "Перенесено эмбрионов"
COL_OUTCOME  = "Исход переноса"
COL_PRAI     = "Predictions"

N_SIM        = 2000
SEED_BASE    = 42
DEFAULT_BMI  = 23.0
DEFAULT_AMH  = 2.0
DEFAULT_AFC  = 12
SPERM_SOURCE = "ejaculate"

OPU_COLUMNS = [
    "Patient Full Name","DOB","ID","Date OPU","Date ET","BMI","AMH",
    "Attempt","AFC","Age","N folicules OPU","OCC","MII","Inseminated",
    "2pN","Cleavage","Bl","Good Bl","Cryo","ET","Day of ET","Preg","PRAI","DIGITAL TWIN",
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_age(v):
    import re
    if v is None: return None
    if isinstance(v, (int, float)):
        x = float(v)
        return x if 10 < x < 60 else None
    m = re.search(r"(\d+)", str(v).strip())
    if m:
        x = float(m.group(1))
        return x if 10 < x < 60 else None
    return None

def num(v, cast=float):
    if v is None: return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
        if v in ("", "-", "—", "н/д", "NA", "nan", "None"): return None
    try:
        x = float(v)
        if np.isnan(x): return None
        return int(round(x)) if cast is int else x
    except: return None

def opt_int(v):
    iv = num(v, int)
    return iv if (iv is not None and iv > 0) else None

def binarize_outcome(val):
    if pd.isna(val) or str(val).strip() in ("", "0", "нет", "Нет"): return np.nan
    s = str(val).lower().strip()
    if "не наступила" in s: return 0
    if "беременность" in s or "клиническая" in s or "биохимическая" in s: return 1
    if s in ("да", "yes", "1"): return 1
    if s in ("нет", "no"):      return 0
    return np.nan

def norm_key(s):
    return str(s).strip().upper() if pd.notna(s) and str(s).strip() else None


# ── Main batch runner ─────────────────────────────────────────────────────────
def run_dt_batch(records, clinic_name, analytics_csv):
    """
    Run DT pipeline for each patient.
    - Writes full 66-col rows to dt_predictions.csv via save_analytics_record
    - Returns maps for post-merge, including the L7 DIGITAL TWIN headline
    """
    from ivf_core import predict_single_patient, save_analytics_record

    os.makedirs(os.path.dirname(analytics_csv), exist_ok=True)

    opu_rows    = []
    prai_map    = {}   # norm_key(patient_id) -> float
    preg_map    = {}   # norm_key(patient_id) -> 0/1/NaN
    dt_map      = {}   # norm_key(patient_id) -> BEFE L7 posterior
    outcome_map = {}   # norm_key(patient_id) -> str

    ok = skip = err = 0

    for i, rec in enumerate(records):
        fio = str(rec.get(COL_FIO, "") or "").strip()
        if fio.lower() in ("итого", "") or not fio:
            skip += 1
            continue

        age = parse_age(rec.get(COL_AGE))
        if age is None:
            skip += 1
            continue

        pid      = str(rec.get(COL_ID, "") or "")
        attempt  = opt_int(rec.get(COL_ATTEMPT)) or 1
        foll     = opt_int(rec.get(COL_FOLL))
        okk      = opt_int(rec.get(COL_OKK))
        insem    = opt_int(rec.get(COL_INSEM))
        real_mii = opt_int(rec.get(COL_MII))

        prai_raw = rec.get(COL_PRAI)
        try:
            prai_val = float(prai_raw) if prai_raw not in (None, "", 0, "0") else None
        except:
            prai_val = None

        real_pn2    = opt_int(rec.get(COL_PN2))
        real_cleav  = opt_int(rec.get(COL_CLEAV))
        real_bl     = opt_int(rec.get(COL_BL))
        real_goodbl = opt_int(rec.get(COL_GOODBL))
        real_cryo   = opt_int(rec.get(COL_CRYO))
        real_outcome= str(rec.get(COL_OUTCOME) or "").strip()
        preg_binary = binarize_outcome(real_outcome)

        preg_flag = "Да" if preg_binary == 1 else ("Нет" if preg_binary == 0 else "")

        # Store for post-merge
        key = norm_key(pid)
        if key:
            prai_map[key]    = prai_val
            preg_map[key]    = preg_binary
            outcome_map[key] = real_outcome

        try:
            result = predict_single_patient(
                age=age, amh=DEFAULT_AMH, afc=DEFAULT_AFC, bmi=DEFAULT_BMI,
                attempt=attempt, follicles=foll, sperm_source=SPERM_SOURCE,
                known_okk=okk, known_mii=insem,
                n_sim=N_SIM, seed=SEED_BASE + i,
            )
            res = result["res"]

            def r4(v):
                try: return round(float(v), 4)
                except: return None

            p_l7, _, _ = compute_l7_posterior(
                result,
                age=age, amh=DEFAULT_AMH, afc=DEFAULT_AFC, bmi=DEFAULT_BMI,
                base_dir=_HERE,
            )
            dt_headline = r4(p_l7) if p_l7 is not None else r4(res.get("p_overall_cycle"))
            if key:
                dt_map[key] = dt_headline

            # ── Write full 66-col row to dt_predictions.csv ───────────────
            save_analytics_record(
                result=result, age=age, amh=DEFAULT_AMH, afc=DEFAULT_AFC,
                bmi=DEFAULT_BMI, attempt=attempt, sperm_source=SPERM_SOURCE,
                follicles=foll, clinic_name=clinic_name,
                patient_name=fio, patient_id=pid,
                real_pn2=real_pn2 or "", real_cleav=real_cleav or "",
                real_bl=real_bl or "", real_goodbl=real_goodbl or "",
                real_cryo=real_cryo or "", real_outcome=real_outcome,
                notes=f"batch;prai={'ok' if prai_val else 'none'}",
                analytics_csv=analytics_csv,
            )

            # ── OPU table row ─────────────────────────────────────────────
            opu_rows.append({
                "Patient Full Name": fio,
                "DOB":               rec.get(COL_DOB, ""),
                "ID":                pid,
                "Date OPU":          rec.get(COL_DATE_OPU, ""),
                "Date ET":           rec.get(COL_DATE_ET, ""),
                "BMI":               DEFAULT_BMI,
                "AMH":               num(rec.get(COL_AMH), float) or "",
                "Attempt":           attempt,
                "AFC":               num(rec.get(COL_AFC), int) or "",
                "Age":               age,
                "N folicules OPU":   foll or "",
                "OCC":               okk or "",
                "MII":               real_mii or "",
                "Inseminated":       insem or "",
                "2pN":               real_pn2 or "",
                "Cleavage":          real_cleav or "",
                "Bl":                real_bl or "",
                "Good Bl":           real_goodbl or "",
                "Cryo":              real_cryo or "",
                "ET":                num(rec.get(COL_ET_N), int) or "",
                "Day of ET":         num(rec.get(COL_DAY_ET), int) or "",
                "Preg":              preg_flag,
                "PRAI":              r4(prai_val) if prai_val else "",
                "DIGITAL TWIN":      dt_headline,
            })

            ok += 1
            nn_ok = "N" if result["nn_available"]   else "-"
            cs_ok = "C" if result["csdi_available"] else "-"
            gn_ok = "G" if result["gnn_available"]  else "-"
            prai_str = f"{prai_val:.3f}" if prai_val is not None else "n/a"
            print(f"  [{i+1:03d}] {fio[:42]:<42}  age={age:.0f}  "
                  f"p={dt_headline}  {nn_ok}{cs_ok}{gn_ok}  prai={prai_str}")

        except Exception as exc:
            print(f"  [{i+1:03d}] ERROR: {fio} | {exc}")
            err += 1

    print(f"\n  OK={ok}  skipped={skip}  errors={err}")
    return opu_rows, prai_map, preg_map, dt_map, outcome_map


# ── OPU xlsx writer ───────────────────────────────────────────────────────────
def write_opu_xlsx(opu_rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OPU_DT"
    ws.append(OPU_COLUMNS)
    hfont = Font(name="Arial", bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1B4F72")
    for col in range(1, len(OPU_COLUMNS)+1):
        c = ws.cell(row=1, column=col)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    prai_fill = PatternFill("solid", start_color="E8F5E9")
    dt_fill   = PatternFill("solid", start_color="FCE4D6")
    for row_data in opu_rows:
        ws.append([row_data.get(k, "") for k in OPU_COLUMNS])
    prai_col = OPU_COLUMNS.index("PRAI") + 1
    dt_col   = OPU_COLUMNS.index("DIGITAL TWIN") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=prai_col).fill = prai_fill
        ws.cell(row=row, column=dt_col).fill   = dt_fill
    for col in range(1, len(OPU_COLUMNS)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
    ws.freeze_panes = "A2"
    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("input",  help="Predictions xlsx")
    parser.add_argument("output", nargs="?", default=".", help="Output dir")
    parser.add_argument("--clinic", default="", help="Clinic name")
    args = parser.parse_args()

    out_dir = args.output
    os.makedirs(out_dir, exist_ok=True)

    analytics_csv = os.path.join(_HERE, "dt_analytics_data", "dt_predictions.csv")

    print(f"\n{'='*60}")
    print(f"  IVF Digital Twin v6.2 -- Batch + Postprocess")
    print(f"  Input : {args.input}")
    print(f"  Clinic: {args.clinic or '(not set)'}")
    print(f"{'='*60}\n")

    # ── Read predictions file ─────────────────────────────────────
    import openpyxl
    wb  = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    ws  = wb.active
    rows= list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    recs= []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        recs.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r)))})

    prai_count = sum(1 for r in recs if r.get(COL_PRAI) not in (None, "", 0, "0"))
    print(f"Records loaded: {len(recs)}  |  PRAI filled: {prai_count}\n")

    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # ── Step 1: run DT pipeline, write dt_predictions.csv ────────
    print("[1/3] Running DT pipeline...\n")
    opu_rows, prai_map, preg_map, dt_map, outcome_map = run_dt_batch(
        recs, args.clinic, analytics_csv
    )

    # ── Step 2: write OPU xlsx ────────────────────────────────────
    opu_path = os.path.join(out_dir, f"OPU_table_filled_{ts}.xlsx")
    write_opu_xlsx(opu_rows, opu_path)
    print(f"\n[2/3] OPU table -> {opu_path}")

    # ── Step 3: build dt_analytics_ready from dt_predictions.csv ─
    # Read the full 66-col CSV just written, then add PRAI + Preg
    print(f"[3/3] Building analytics files from dt_predictions.csv...")

    df = pd.read_csv(analytics_csv)

    # Normalise key for merge
    df['_key'] = df['patient_id'].astype(str).str.strip().str.upper()

    # Add PRAI, Preg, real_outcome from maps
    df['PRAI']         = df['_key'].map(prai_map)
    df['Preg']         = df['_key'].map(preg_map)
    df['DIGITAL TWIN'] = df['_key'].map(dt_map)
    df['real_outcome'] = df['_key'].map(outcome_map).fillna(df.get('real_outcome', ''))

    # Deduplicate: keep last run per patient (consistent with app.py behaviour)
    if 'timestamp' in df.columns:
        df = (df.sort_values('timestamp', ascending=True)
                .drop_duplicates(subset=['_key'], keep='last')
                .copy())

    # OPU-name aliases so existing Datalore code works without changes
    # Алиасы: добавляем OPU-имена рядом с DT-именами (не переименовываем)
    # known_mii -> и MII и Inseminated (две разные колонки)
    alias_list = [
        ('age',            'Age'),
        ('attempt_number', 'Attempt'),
        ('follicles_tvp',  'N folicules OPU'),
        ('known_okk',      'OCC'),
        ('known_mii',      'MII'),           # отдельно
        ('known_mii',      'Inseminated'),   # отдельно — список, не словарь!
        ('med_blasts',     'med_bl'),
        ('med_good',       'med_goodbl'),
        ('real_pn2',       '2pN'),
        ('real_bl',        'Bl'),
        ('real_goodbl',    'Good Bl'),
    ]
    for dt_col, opu_col in alias_list:
        if dt_col in df.columns and opu_col not in df.columns:
            df[opu_col] = df[dt_col]

    # Fill OCC/MII/N folicules from Inseminated if still NaN
    for col in ['N folicules OPU', 'OCC', 'MII']:
        if col in df.columns:
            df[col] = df[col].fillna(df.get('Inseminated', df.get('known_mii')))

    # ── Write outputs ─────────────────────────────────────────────
    ready_path = os.path.join(out_dir, f"dt_analytics_ready_{ts}.csv")
    df.to_csv(ready_path, index=False, encoding="utf-8-sig")
    print(f"     dt_analytics_ready       -> {ready_path}")
    print(f"     Columns: {df.shape[1]}  Rows: {len(df)}")

    df_preg = df[df['Preg'].notna()].copy()
    df_preg['Preg'] = df_preg['Preg'].astype(int)
    n_pos = int(df_preg['Preg'].sum())
    n_neg = int((df_preg['Preg'] == 0).sum())
    out_path = os.path.join(out_dir, f"dt_analytics_with_outcome_{ts}.csv")
    df_preg.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"     dt_analytics_with_outcome -> {out_path}")
    print(f"     dt_predictions.csv         -> {analytics_csv}")

    # ── Summary ───────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Patients processed : {len(df)}")
    print(f"  With outcome (Preg): {len(df_preg)}  (+{n_pos} / -{n_neg})")
    print(f"  PRAI filled        : {df['PRAI'].notna().sum()}/{len(df)}")
    print(f"  CSV columns        : {df.shape[1]}  (manual app.py = 66+)")
    if len(df_preg) > 0:
        print(f"\n  Model means (n={len(df_preg)}):")
        for col, name in [
            ("p_overall_cycle","DT overall"), ("bayes_mean","Bayes"),
            ("p_kat_raw","KAT L3"), ("p_nvsa","NVSA L3"),
            ("p_csdi","CSDI L5"), ("p_gnn_ens","GNN L6"),
            ("DIGITAL TWIN","BEFE L7"), ("PRAI","PRAI"),
        ]:
            vals = pd.to_numeric(df_preg.get(col, pd.Series()), errors="coerce").dropna()
            if len(vals) > 0:
                print(f"    {name:<14} n={len(vals):3d}  mean={vals.mean():.3f}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
